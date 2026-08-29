#!/usr/bin/env python3
"""Reusable redirect-map tooling for the docs restructure (Option 2).

Commands:
  python scripts/redirects.py rebuild
      Rebuild _docs-ia/testmu-docs-redirects.xlsx from every
      _docs-ia/*-redirect-map.csv, joining 6-month clicks from
      _docs-ia/gsc-pages.csv and assigning a merge tier + target type.
      Whole-page targets with >0 clicks are highlighted (soft-404 risk).

  python scripts/redirects.py add <product> "<source>" "<target>" ["note"]
      Append one redirect to _docs-ia/<product>-redirect-map.csv.
      <source>/<target> may be a full URL, a /support/docs/... path, or a slug.

  python scripts/redirects.py bulk <product> <target-slug> <slug1> <slug2> ...
      Tier-3 bulk redirect: point many old slugs at one overview page.
"""
import csv, glob, os, re, sys
from collections import Counter

BASE = "https://www.testmuai.com"
PFX = "/support/docs/"
IA = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "_docs-ia"))
GSC = os.path.join(IA, "gsc-pages.csv")


def slug_of(url):
    u = url.split("#")[0].split("?")[0]
    m = re.search(r"/support/docs/([^/]+)/?", u)
    return m.group(1) if m else u.strip().strip("/").split("/")[-1]


def as_url(x):
    if x.startswith("http"):
        return x
    if x.startswith("/"):
        return BASE + x
    return BASE + PFX + x.strip("/") + "/"


def load_clicks():
    clicks = {}
    if not os.path.exists(GSC):
        print(f"WARN: {GSC} not found; clicks will be 0")
        return clicks
    with open(GSC, encoding="utf-8-sig", newline="") as fh:
        r = csv.reader(fh)
        next(r, None)  # header
        for row in r:
            if len(row) < 2:
                continue
            url = row[0]
            if "#" in url or "?" in url:  # page-level rows only
                continue
            try:
                clicks[slug_of(url)] = int(str(row[1]).replace(",", "").strip())
            except ValueError:
                continue
    return clicks


def tier(c):
    if c >= 101:
        return "Tier 1 (Core 101+)"
    if c >= 1:
        return "Tier 2 (Long-tail 1-100)"
    return "Tier 3 (Zero)"


def ttype(t):
    if "?framework=" in t:
        return "Framework tab"
    if "#" in t:
        return "Anchored section"
    return "Whole page"


def read_map(f):
    out = []
    with open(f, encoding="utf-8-sig", newline="") as fh:
        for d in csv.DictReader(fh):
            src = (d.get("Source URL") or d.get("Source URL (old)") or "").strip()
            tgt = (d.get("Target URL") or d.get("Target URL (301)") or "").strip()
            if src and tgt:
                out.append((src, tgt))
    return out


def rebuild():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    clicks = load_clicks()
    rows = []
    for f in sorted(glob.glob(os.path.join(IA, "*-redirect-map.csv"))):
        product = os.path.basename(f).replace("-redirect-map.csv", "")
        for src, tgt in read_map(f):
            c = clicks.get(slug_of(src), 0)
            rows.append([src, tgt, c, tier(c), ttype(tgt), product])
    rows.sort(key=lambda r: -r[2])

    wb = Workbook(); ws = wb.active; ws.title = "Redirects"
    ws.append(["URL to be Redirected", "Target URL", "6-Month Clicks", "Tier", "Target Type", "Product"])
    fill = PatternFill("solid", fgColor="4472C4")
    warn = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9"); bd = Border(thin, thin, thin, thin)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill; c.border = bd
    for r in rows:
        ws.append(r)
        cells = ws[ws.max_row]
        for c in cells:
            c.border = bd; c.alignment = Alignment(vertical="center")
        if r[2] >= 1 and r[4] == "Whole page":
            for c in cells:
                c.fill = warn
    for i, w in enumerate([74, 82, 14, 24, 18, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    out = os.path.join(IA, "testmu-docs-redirects.xlsx")
    wb.save(out)
    print(f"Rebuilt {out}: {len(rows)} redirects")
    for k, v in sorted(Counter(r[3] for r in rows).items()):
        print(f"  {k}: {v}")
    flagged = [r for r in rows if r[2] >= 1 and r[4] == "Whole page"]
    if flagged:
        print(f"  soft-404 review (whole-page targets, >0 clicks): {len(flagged)}")


def _append(product, src, tgt, note):
    f = os.path.join(IA, f"{product}-redirect-map.csv")
    newfile = not os.path.exists(f)
    with open(f, "a", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        if newfile:
            w.writerow(["Source URL", "Target URL", "Type", "Notes"])
        w.writerow([as_url(src), as_url(tgt), "Redirect", note])
    print(f"  + {as_url(src)} -> {as_url(tgt)}")


def add(product, source, target, note=""):
    _append(product, source, target, note)
    print(f"Added 1 redirect to {product}-redirect-map.csv")


def bulk(product, target_slug, *slugs):
    for s in slugs:
        _append(product, s, target_slug, "Tier-3 bulk redirect (0 clicks)")
    print(f"Added {len(slugs)} bulk redirects to {product}-redirect-map.csv")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "rebuild"
    if cmd == "rebuild":
        rebuild()
    elif cmd == "add":
        add(*sys.argv[2:])
    elif cmd == "bulk":
        bulk(*sys.argv[2:])
    else:
        print(__doc__)
